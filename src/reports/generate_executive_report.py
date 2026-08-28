import os
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_executive_package():
    os.makedirs("reports", exist_ok=True)
    
    # 1. Load Processed Datasets
    orders_df = pd.read_csv("data/processed/cleaned_logistics_data.csv")
    products_df = pd.read_csv("data/raw/product_master.csv")
    
    # 2. Compute Aggregations
    # Supplier Performance KPI Table
    supplier_kpis = orders_df.groupby("supplier").agg(
        total_pos=("po_number", "count"),
        total_spend=("total_cost", "sum"),
        avg_lead_time=("actual_lead_time_days", "mean"),
        avg_variance=("lead_time_variance", "mean"),
        on_time_count=("is_on_time", lambda x: (x == True).sum())
    ).reset_index()
    
    supplier_kpis["on_time_rate_%"] = (supplier_kpis["on_time_count"] / supplier_kpis["total_pos"]) * 100
    supplier_kpis = supplier_kpis.drop(columns=["on_time_count"]).sort_values(by="total_spend", ascending=False)

    # SKU Inventory & Replenishment Policy Table
    lead_stats = orders_df.groupby("sku").agg(
        avg_lead=("actual_lead_time_days", "mean"),
        std_lead=("actual_lead_time_days", "std"),
        total_volume=("order_quantity", "sum")
    ).reset_index()
    
    inv_summary = pd.merge(products_df, lead_stats, on="sku")
    inv_summary["avg_daily_demand"] = inv_summary["total_volume"] / 365.0
    z_score = 1.65 # 95% service level
    
    inv_summary["dyn_safety_stock"] = np.ceil(z_score * inv_summary["std_lead"].fillna(0) * inv_summary["avg_daily_demand"]).astype(int)
    inv_summary["reorder_point"] = np.ceil((inv_summary["avg_daily_demand"] * inv_summary["avg_lead"]) + inv_summary["dyn_safety_stock"]).astype(int)
    inv_summary["working_capital_diff_$"] = (inv_summary["safety_stock"] - inv_summary["dyn_safety_stock"]) * inv_summary["unit_cost"]

    # Actionable Alerts (Delayed Shipments)
    actionable_orders = orders_df[orders_df["po_status"].isin(["Delayed", "Cancelled"])][[
        "po_number", "sku", "category", "supplier", "order_date", 
        "expected_lead_time_days", "actual_lead_time_days", "lead_time_variance", "po_status", "total_cost"
    ]].sort_values(by="lead_time_variance", ascending=False)

    # 3. Write to Multi-Sheet Excel Workbook
    report_path = "reports/ESAB_Logistics_Executive_Report.xlsx"
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        supplier_kpis.to_excel(writer, sheet_name="Supplier Performance", index=False)
        inv_summary.to_excel(writer, sheet_name="Inventory Policy & ROP", index=False)
        actionable_orders.to_excel(writer, sheet_name="Shipment Exceptions", index=False)

    # 4. Format Workbook Aesthetics (Headers, Fills, Borders)
    wb = load_workbook(report_path)
    
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        ws.views.sheetView[0].showGridLines = True
        
        # Style Header Row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 24

        # Style Data Cells & Auto-Fit Column Widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                cell.border = thin_border
                val_str = str(cell.value or '')
                max_len = max(max_len, len(val_str))
                
                if cell.row > 1 and isinstance(cell.value, (int, float)):
                    if "cost" in str(ws.cell(row=1, column=cell.column).value).lower() or "$" in str(ws.cell(row=1, column=cell.column).value).lower() or "spend" in str(ws.cell(row=1, column=cell.column).value).lower():
                        cell.number_format = '$#,##0.00'
                    elif "rate" in str(ws.cell(row=1, column=cell.column).value).lower() or "%" in str(ws.cell(row=1, column=cell.column).value).lower():
                        cell.number_format = '0.0"%"'
                    elif isinstance(cell.value, float):
                        cell.number_format = '#,##0.00'
            
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(report_path)
    print(f"Executive Report successfully generated and styled: {report_path}")

if __name__ == "__main__":
    build_executive_package()